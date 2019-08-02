import sys
import os
import time
import codecs

try:
    import argparse
    import urllib.error
    from Bio import SeqIO
    from Bio import Entrez
    from config import config
    from fuzzywuzzy import fuzz
    from fuzzywuzzy import process
    from openpyxl import Workbook
    from openpyxl import load_workbook
except ImportError as err:
    print("Import error!: " + err.name)
    exit(1)


Entrez.email = "bob@hotmail.com"


# Parsing for standard input arguments that user might enter for modularity
parser = argparse.ArgumentParser()
parser.add_argument('accessions')
parser.add_argument("--accessions", help="File Containing list of Accession IDs")
parser.add_argument("--recognition", help="Custom Recognition List")
parser.add_argument("--feature", help="Custom Feature List")
parser.add_argument("--header", help="Custom Headers List and data to pull out")
parser.add_argument('-f', action="store_true", help="Flag to produce Fasta file")
parser.add_argument('-t', action="store_true", help="Flag to produce TSV file")
parser.add_argument('-l', action="store_true", help="Flag to produce Logs file")
args = parser.parse_args()


def fetch_sequence(feature, recognition_list, record, feature_list):
    for val in feature.qualifiers.items():
        gene = []
        sequence = []
        location = []
        qualifiers = []
        pseudo = []
        for value in val:
            if isinstance(value, list):
                value = value[0]

            if value in recognition_list and (feature.type in feature_list):
                try:
                    gene.append(value)
                    sequence.append(feature.location.extract(record).seq)
                    location.append(feature.location)

                    for qualifier in feature.qualifiers.items():
                        key = qualifier[0]
                        value = qualifier[1]
                        if isinstance(value, list):
                            value = qualifier[1][0]
                        qualifier_string = str(key) + ": " + str(value)
                        qualifiers.append(qualifier_string)

                    for item in feature.qualifiers:
                        matchRatio = fuzz.ratio(item, "pseudo")
                        if matchRatio > 80:
                            pseudo.append(item)
                    return feature.type, gene, qualifiers, location, sequence, pseudo
                except KeyError:
                    pass
        return feature.type, gene, qualifiers, location, sequence, pseudo


def fetch_match(header, record):
    """
    Function to try to retrieve an annotation from a SeqRecord object, return empty string if not found

    :param header: header is the header that we're looking to match and extract     -REQUIRED
    :param record: record is the genbank record for a specific accession ID         -REQUIRED
    :return:
    """
    output = []
    # Try and see if the header can be found in the root of the object record
    try:
        methods = dir(record)
        for method in methods:
            matchRatio = fuzz.ratio(header, method)
            if matchRatio > 70:
                method_called = getattr(record, method)
                output.append(method_called)
        if len(output) != 0:
            return output
        pass
    except KeyError:
        pass
    try:
        unPackKeys = [*record.annotations]
        for key in unPackKeys:
            matchRatio = fuzz.ratio(header, key)
            if matchRatio > 70:
                if isinstance(record.annotations[key], list):
                    for item in record.annotations[key]:
                        string = str(item).rstrip().replace('\n', '| ')
                        # print("|" + string + "|")
                        output.append(string)
        if len(output) != 0:
            return output
        pass
    except KeyError:
        pass
    # Finally, if the header doesn't exist in the root object annotations, check the associated objects.
    try:
        for featureIndex in range(0, len(record.features)):
            try:
                feature_dict = record.features[featureIndex]
                unPackKeys = [*feature_dict.qualifiers]
                for key in unPackKeys:
                    matchRatio = fuzz.ratio(header, key)
                    if matchRatio > 70:
                        output.append(feature_dict.qualifiers[key])
            except KeyError:
                pass
        return output
    except:
        return output


def check_standard_input_files(stdin_file, stdin_list):
    if stdin_file is not None:
        try:
            lines = open(stdin_file).readlines()
            for line in lines:
                line = line.rstrip('\n')
                line = line.strip()
                stdin_list.append(line)
        except FileNotFoundError:
            print("Error: File was not found: " + stdin_file)
            exit(1)
    else:
        return 1


def grouper(iterable, n):
    for i in range(0, len(iterable), n):
        yield iterable[i:i + n]


def listToString(candidate):
    return sum(([candidate] if not isinstance(candidate, list) else listToString(candidate) for candidate in candidate), [])


def load_execute():
    """
    Function that loads all user arguments from console and processes/assigns them to variables
    setting up the environment for data execution

    To run the program, only the input file name is required. However other arguments can be input to add modularity
    """
    # If no arguments passed, exit
    if not len(sys.argv) > 1:
        print("Error: No Arguments passed")
        exit(1)

    # list of accessions that need be processed
    if len(sys.argv) == 1:
        input_file_name = sys.argv[1]
    else:
        input_file_name = args.accessions

    if not os.path.isfile(input_file_name):
        print("File could not be found: " + input_file_name)

    # the custom recognition files
    recognition_file = args.recognition
    feature_file = args.feature
    header_file = args.header

    # list to store input data
    recognition_list = []
    feature_list = []
    header_list = []

    ret = check_standard_input_files(recognition_file, recognition_list)
    if ret == 1 or (len(recognition_list) < 1):
        recognition_list = config.default_rec_list

    ret = check_standard_input_files(feature_file, feature_list)
    if ret == 1 or (len(feature_list) < 1):
        feature_list = config.default_fet_list

    ret = check_standard_input_files(header_file, header_list)
    if ret == 1 or (len(header_list) < 1):
        header_list = config.default_head_list

    input_file_name_edit = str(os.path.basename(input_file_name).split(".")[0])
    if args.f is True:
        fasta_file = str("Result_" + input_file_name_edit + ".fasta")
    else:
        fasta_file = None

    if args.t is True:
        tsv_file = str("tsv_" + input_file_name_edit + ".tsv")
    else:
        tsv_file = None

    if args.l is True:
        log_file = str("log_" + input_file_name_edit + ".txt")
    else:
        log_file = None

    execute(input_file_name, fasta_file, tsv_file, log_file, header_list, feature_list, recognition_list)


def execute(input_file_name, fasta_file, tsv_file, log_file, header_list, feature_list, recognition_list):
    """
    Main Processing function that parses that downloads and parses the information

    :param input_file_name: list of accession IDs                                   - REQUIRED
    :param fasta_file: Name of the Fasta file the user wants to set                 - NOT REQUIRED
    :param tsv_file: Name of the tsv file the user wants to save data in            - NOT REQUIRED
    :param log_file: Name of the log file the user wants to produce                 - NOT REQUIRED
    :param header_list: Header list                                                 - REQUIRED
    :param feature_list: Feature list to look for "Hooks" to extract Sequences      - NOT REQUIRED UNLESS RECOGNITION LIST EXISTS
    :param recognition_list: Recognition "Hook" list                                - NOT REQUIRED UNLESS FEATURE LIST EXISTS
    :return: Nothing
    """

    # Count the number of Chunks processed
    counter = 0
    # delay counter for HTTP errors
    long_delay = 0

    # Logs list. Accessions with no Sequences
    logs_list = []

    # Fasta List and sequence List
    fasta_list = []
    fasta_feature = []
    fasta_sequence = []

    # Create workbook without creating it on file
    wb = Workbook()
    # Select worksheet
    ws = wb.active
    # Populate the sheet's headers with user inserted headers
    for index, header in enumerate(header_list):
        ws.cell(row=1, column=index+1, value=header)

    numberHeaders = len(header_list)
    # feature.type, gene, qualifiers, location, sequence, pseudo
    ws.cell(row=1, column=numberHeaders + 1, value="Recognition List")
    ws.cell(row=1, column=numberHeaders + 2, value="Qualifiers")
    ws.cell(row=1, column=numberHeaders + 3, value="Location")
    ws.cell(row=1, column=numberHeaders + 4, value="sequence")
    ws.cell(row=1, column=numberHeaders + 5, value="Pseudo")

    with codecs.open(input_file_name, 'r') as infile:
        chunks = grouper(infile.read().split("\n"), 300)
        for chunk in chunks:
            print("Processed " + str(counter) + " records...")
            while "" in chunk:
                chunk.remove('')
            acc_list = ",".join(chunk)
            try:
                fetched_gb = Entrez.efetch(db="nucleotide", id=acc_list, rettype="gbwithparts", retmode="text")
                for index, record in enumerate(SeqIO.parse(fetched_gb, "gb")):
                    print("Processing: " + record.id)
                    keyValues = []
                    # First get data to populate header fields
                    for header in header_list:
                        found = fetch_match(header, record)
                        keyValues.append(found)

                    seqBatch = []
                    # Fetch the sequence data
                    for feature in record.features:
                        seqInfo = fetch_sequence(feature, recognition_list, record, feature_list)
                        # If no sequence exists don't append to the list
                        if len(seqInfo[1]) != 0:
                            seqBatch.append(seqInfo)

                    numberRowsToPrint = len(seqBatch)

                    # TODO: Patch Remove Duplicates process
                    # if numberRowsToPrint > 1:
                    #     # Remove Duplicates
                    #     visited = set()
                    #     Output = []
                    #     for feature, gene, sequence, valueKey in seqBatch:
                    #         if not sequence[0] in visited:
                    #             visited.add(sequence[0])
                    #             Output.append((feature, gene, sequence, valueKey))
                    #     seqBatch = Output
                    #     numberRowsToPrint = len(seqBatch)

                    if numberRowsToPrint == 0:
                        numberRowsToPrint = 1

                    # print(seqBatch)
                    # Next populate the fields
                    for relativeLine in range(numberRowsToPrint):
                        workingRow = ws.max_row
                        for indexHeader in range(len(keyValues) + 1):
                            if indexHeader < len(keyValues):
                                extractValue = listToString(keyValues[indexHeader])
                                if len(extractValue) == 0:
                                    extractValue = " "
                                else:
                                    extractValue = ", ".join(str(x) for x in extractValue)
                                    ws.cell(row=workingRow+1, column=indexHeader + 1, value=extractValue)
                            else:
                                # If no sequences were found here. Add to logs list
                                if len(seqBatch) != 0:
                                    try:
                                        # Recognition
                                        ws.cell(row=workingRow+1, column = indexHeader + 1, value=str(seqBatch[relativeLine][1][0]))
                                        # Qualifiers found
                                        join_string = ", ".join(str(x) for x in seqBatch[relativeLine][2])
                                        ws.cell(row=workingRow+1, column=indexHeader + 2, value=str(join_string))
                                        # Location
                                        ws.cell(row=workingRow + 1, column=indexHeader + 3, value=str(seqBatch[relativeLine][3][0]))
                                        # Sequence
                                        ws.cell(row=workingRow + 1, column=indexHeader + 4, value=str(seqBatch[relativeLine][4][0]))
                                        # IsPseudo
                                        if len(seqBatch[relativeLine][5]) == 0:
                                            ws.cell(row=workingRow + 1, column=indexHeader + 5, value="")
                                        else:
                                            ws.cell(row=workingRow + 1, column=indexHeader + 5,value=str(seqBatch[relativeLine][5][0]))

                                        fasta_list.append(record.id)
                                        fasta_feature.append(seqBatch[relativeLine][1][0])
                                        fasta_sequence.append(seqBatch[relativeLine][4][0])
                                    except Exception as e:
                                        print(e)
                                        print("Relative Line: " + str(relativeLine))
                                        print("Number of Repeats: " + str(numberRowsToPrint))
                                        exit(1)
                                else:
                                    logs_list.append(record.id)
            except urllib.error.HTTPError:
                if urllib.error.HTTPError.code == 429:
                    time.sleep(5)
                    pass
                else:
                    a = False
            long_delay = long_delay + 1
            if long_delay == 20:
                time.sleep(20)
                long_delay = 0

    # Revisit the headers to find lat_lon
    for index, header in enumerate(header_list):
        if header == 'lat_lon':
            ws.insert_cols(index + 2)   # Lat
            ws.insert_cols(index + 3)   # Lon
            ws.cell(row=1, column=index + 2, value="Lat")
            ws.cell(row=1, column=index + 3, value="Lon")
            for row in range(2, ws.max_row+1):
                lat_lon = ws.cell(row=row, column=index+1).value
                lat_lon = lat_lon.split(" ")
                lat = lat_lon[0] + " " + lat_lon[1]
                lon = lat_lon[2] + " " + lat_lon[3]
                ws.cell(row=row, column=index + 2, value= lat)
                ws.cell(row=row, column=index + 3, value=lon)

    if fasta_file is not None:
        fasta_file_ptr = open(fasta_file, "w")
        for index in range(len(fasta_list)):
            fasta_file_ptr.write("> " + str(fasta_list[index]) + "|" + str(fasta_feature[index]) + "\n" + str(fasta_sequence[index]) + "\n")
        fasta_file_ptr.close()

    if tsv_file is not None:
        tsv_file_ptr = open(tsv_file, "w")
        max_row = ws.max_row
        max_col = ws.max_column
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                text = ws.cell(row=row, column=col).value
                if text is None:
                    text = " "
                tsv_file_ptr.write(text)
                tsv_file_ptr.write("\t")
            tsv_file_ptr.write("\n")
        tsv_file_ptr.close()

    if log_file is not None:
        log_file_ptr = open(log_file, "w")
        for accessionID in logs_list:
            log_file_ptr.write(accessionID + "\t" + "No Sequence Found!" + "\n")
        log_file_ptr.close()

    # Save the Workbook that's been created in Memory!
    wb.save("genbankParsedData.xlsx")


if __name__ == '__main__':
    load_execute()
